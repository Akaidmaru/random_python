from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import date

# This creates the pattern i want the cell to be filled in
fill_pattern = PatternFill(
	fill_type='solid',
	start_color='00FFFF',
	end_color='00FFFF'
	)

# This creates alignment
alignment_center = Alignment(
	horizontal='center',
	vertical='center'
	)

# This creates the bolding
font_bold = Font(
	bold=True
	)

wb = Workbook()
ws = wb.active

# I open .txt and read it
with open('C:/users/IA/Desktop/Littleton/CALLS.txt', encoding= 'utf8') as f:
	lines = f.readlines()
	
	# Create a counter using enumerate so i can read each line and also split each line with a whitepase to get lists
	for count, line in enumerate(lines, start=1):
		new_list = line.split(' ')

		# This creates the header of the xlsx.
		if len(new_list) == 1:
			ws.cell(row=count, column=1).value = 'DATE'
			ws.cell(row=count, column=2).value = 'START CALL TIME'
			ws.cell(row=count, column=3).value = 'END CALL TIME'
			ws.cell(row=count, column=4).value = 'TELEPHONE'
			ws.cell(row=count, column=5).value = 'CLAIM NO'
			ws.cell(row=count, column=6).value = 'EX/NEW'
			ws.cell(row=count, column=7).value = 'NOTES'

		# If the lenght of the list splitted is == 3.
		elif len(new_list) == 3:
			if 'EXT' in new_list[1]:
				ws.cell(row=count, column=2).value = new_list[0]
				ws.cell(row=count, column=2).alignment = alignment_center

				ws.cell(row=count, column=3).value = new_list[-1]
				ws.cell(row=count, column=3).alignment = alignment_center

				ws.cell(row=count, column=4).value = new_list[1]
				ws.cell(row=count, column=4).alignment = alignment_center

				ws.cell(row=count, column=5).value = '-'
				ws.cell(row=count, column=5).alignment = alignment_center

				ws.cell(row=count, column=6).value = '-'
				ws.cell(row=count, column=6).alignment = alignment_center

				ws.cell(row=count, column=7).value = 'INTERPRETING'
				ws.cell(row=count, column=7).alignment = alignment_center

			else:
				ws.cell(row=count, column=2).value = new_list[0]
				ws.cell(row=count, column=2).alignment = alignment_center

				ws.cell(row=count, column=3).value = new_list[-1]
				ws.cell(row=count, column=3).alignment = alignment_center

				ws.cell(row=count, column=4).value = new_list[1]
				ws.cell(row=count, column=4).alignment = alignment_center

				ws.cell(row=count, column=5).value = '-'
				ws.cell(row=count, column=5).alignment = alignment_center

				ws.cell(row=count, column=6).value = '-'
				ws.cell(row=count, column=6).alignment = alignment_center

				ws.cell(row=count, column=7).value = 'INFO'
				ws.cell(row=count, column=7).alignment = alignment_center

		
		elif len(new_list) == 4:
			ws.cell(row=count, column=2).value = new_list[0]
			ws.cell(row=count, column=2).alignment = alignment_center

			ws.cell(row=count, column=3).value = new_list[-1]
			ws.cell(row=count, column=3).alignment = alignment_center

			ws.cell(row=count, column=4).value = new_list[1]
			ws.cell(row=count, column=4).alignment = alignment_center

			ws.cell(row=count, column=5).value = new_list[2]
			ws.cell(row=count, column=5).alignment = alignment_center

			ws.cell(row=count, column=6).value = 'NEW'
			ws.cell(row=count, column=6).alignment = alignment_center

			ws.cell(row=count, column=7).value = 'CLAIM'
			ws.cell(row=count, column=7).alignment = alignment_center

		else:
			ws.cell(row=count, column=2).value = new_list[0]
			ws.cell(row=count, column=2).alignment = alignment_center

			ws.cell(row=count, column=3).value = new_list[-1]
			ws.cell(row=count, column=3).alignment = alignment_center

			ws.cell(row=count, column=4).value = new_list[1]
			ws.cell(row=count, column=4).alignment = alignment_center

			ws.cell(row=count, column=5).value = new_list[2]
			ws.cell(row=count, column=5).alignment = alignment_center

			ws.cell(row=count, column=6).value = 'EX'
			ws.cell(row=count, column=6).alignment = alignment_center

			ws.cell(row=count, column=7).value = 'UPDATE'
			ws.cell(row=count, column=7).alignment = alignment_center

	# Date
	today = date.today()

	ws['A2'] = today.strftime("%m/%d/%Y")
	print(today.strftime("%m/%d/%Y"))

	# Filling
	ws['A1'].fill = fill_pattern
	ws['B1'].fill = fill_pattern
	ws['C1'].fill = fill_pattern
	ws['D1'].fill = fill_pattern
	ws['E1'].fill = fill_pattern
	ws['F1'].fill = fill_pattern
	ws['G1'].fill = fill_pattern
	
	# Alignment
	ws['A1'].alignment = alignment_center
	ws['B1'].alignment = alignment_center
	ws['C1'].alignment = alignment_center
	ws['D1'].alignment = alignment_center
	ws['E1'].alignment = alignment_center
	ws['F1'].alignment = alignment_center
	ws['G1'].alignment = alignment_center

	# Bolding
	ws['A1'].font = font_bold
	ws['B1'].font = font_bold
	ws['C1'].font = font_bold
	ws['D1'].font = font_bold
	ws['E1'].font = font_bold
	ws['F1'].font = font_bold
	ws['G1'].font = font_bold

	# Merging Cells
	ws.merge_cells(start_row=2, start_column=1, end_row=count, end_column=1)
	ws['A2'].alignment = alignment_center

	# Saving Workbook
	wb.save('C:/Users/IA/Desktop/automated_report.xlsx')
